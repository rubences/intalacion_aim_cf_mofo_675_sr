import pandas as pd
from pathlib import Path

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# Datos para la Pestana 1: Hoja de Especificaciones de Componentes
data1 = {
    "ID": ["01", "02", "03", "04", "05", "06"],
    "Sistema / Componente": [
        "AiM EVO4 (Modulo Central)",
        "MyChron3 Dash (Display)",
        "Antena GPS",
        "Potenciometro (Suspension)",
        "Sensor Velocidad (Inductivo)",
        "Mazo de Cables / CAN Bus",
    ],
    "Ubicacion Tecnica": [
        "Subchasis / Bandeja trasera",
        "Cockpit / Arana mecanizada",
        "Colin / Frontal (Cupula)",
        "Horquilla delantera",
        "Soporte pinza freno radial",
        "Ramal principal chasis",
    ],
    "Fijacion / Torque": [
        "Velcro 3M Dual Lock + Bridas",
        "Tornilleria M6 (10 Nm) + Loctite 243",
        "Cinta 3M VHB",
        "Abrazaderas M4 (3-4 Nm)",
        "Tornilleria M6 (10 Nm) + Loctite 243",
        "Cinta de tela (Tesa) / Bridas",
    ],
    "Herramientas / Materiales": [
        "Desengrasante, Nivel de burbuja",
        "Llave dinamometrica, llaves Allen",
        "Alcohol isopropilico para limpieza",
        "Pie de rey, llaves Allen 3mm",
        "Galgas de espesores",
        "Pasacables, funda termorretractil",
    ],
    "Calibracion / Notas Criticas": [
        "Alinear ejes X/Y/Z con la moto. Nivelacion a 0 deg estricta.",
        "Comprobar visibilidad en posicion de acople del piloto.",
        "Sin obstruccion superior (carbono/metal bloquean senal).",
        "Reservar min. 5mm de vastago en compresion maxima.",
        "Gap milimetrico de 1.0 - 2.0 mm respecto al disco.",
        "Aislar de fuentes de interferencia electromagnetica.",
    ],
    "QC (Check)": ["", "", "", "", "", ""],
}

# Datos para la Pestana 2: Protocolo de Instalacion (SOP)
data2 = {
    "Fase": ["1. Prep", "1. Prep", "2. Core", "2. Core", "3. Sens", "3. Sens", "4. Cabl", "5. Soft", "5. Soft"],
    "Tarea Especifica": [
        "Aislamiento electrico",
        "Desmontaje carenados",
        "Montaje de Datalogger",
        "Montaje Dash y GPS",
        "Regulacion Potenciometro",
        "Regulacion Velocidad",
        "Enrutamiento seguro",
        "Setup RaceStudio",
        "Calibracion a Cero",
    ],
    "Especificaciones Tecnicas (El Como)": [
        "Desconectar borne negativo de la bateria.",
        "Retirar plasticos para acceso al chasis.",
        "Instalar EVO4 con tolerancias de vibracion.",
        "Adherir GPS tras desengrasar; atornillar Dash.",
        "Alinear vastago 100% paralelo a la barra.",
        "Ajustar distancia sensor-tornillo del disco.",
        "Guiar mazo principal por el interior del chasis.",
        "Dar contacto, cargar configuracion de 675 SR.",
        "Establecer 0 de suspension extendida.",
    ],
    "Verificacion y Control de Calidad (QC)": [
        "Voltaje 0V confirmado en el sistema.",
        "Tornilleria en bandejas magneticas numeradas.",
        "Nivel confirmado a 0 deg en ejes longitudinal/transversal.",
        "Giro del manillar libre de topes; cables sin tension.",
        "Test fisico: Hundir a tope, verificar que no colapsa.",
        "Giro manual 360 deg: Lectura positiva en todos los tornillos.",
        "Cables alejados >5cm de colectores y sin pinzamiento.",
        "Pilotos LED de EVO4 encendidos, conexion a PC estable.",
        "Telemetria en vivo mostrando valores reales y logicos.",
    ],
    "Tiempo Est.": ["10 min", "30 min", "20 min", "30 min", "45 min", "25 min", "45 min", "20 min", "15 min"],
    "Firma": ["", "", "", "", "", "", "", "", ""],
}

# Datos para la Pestana 3: Checklist Detallado
data3 = {
    "Fase": [
        "Fase 1: Preparacion",
        "Fase 1: Preparacion",
        "Fase 1: Preparacion",
        "Fase 2: Modulo EVO4",
        "Fase 2: Modulo EVO4",
        "Fase 2: Modulo EVO4",
        "Fase 3: Cockpit y GPS",
        "Fase 3: Cockpit y GPS",
        "Fase 3: Cockpit y GPS",
        "Fase 3: Cockpit y GPS",
        "Fase 4: Sensores",
        "Fase 4: Sensores",
        "Fase 4: Sensores",
        "Fase 4: Sensores",
        "Fase 4: Sensores",
        "Fase 5: Enrutado y ECU",
        "Fase 5: Enrutado y ECU",
        "Fase 5: Enrutado y ECU",
        "Fase 6: Config y Pruebas",
        "Fase 6: Config y Pruebas",
        "Fase 6: Config y Pruebas",
        "Fase 6: Config y Pruebas",
        "Fase 6: Config y Pruebas",
    ],
    "Tarea Especifica": [
        "Desconectar la bateria",
        "Desmontaje de carenados",
        "Presentacion en seco",
        "Fijacion del EVO4",
        "Aislamiento de vibraciones",
        "Toma de corriente",
        "Soporte del Dash",
        "Fijacion de la pantalla",
        "Ubicacion de antena GPS",
        "Verificacion de vision",
        "Soportes del potenciometro",
        "Alineacion del potenciometro",
        "PRUEBA CRITICA (Recorrido)",
        "Soporte sensor de velocidad",
        "Ajuste del Gap (Velocidad)",
        "Guiado limpio",
        "Proteccion termica/mecanica",
        "Conexion a ECU",
        "Conectar bateria",
        "Conexion a PC",
        "Configuracion de Moto",
        "Calibracion a Cero",
        "Prueba de giro",
    ],
    "Accion Detallada y Precauciones": [
        "Retirar el borne negativo para evitar cortocircuitos.",
        "Retirar tapas laterales, asiento, colin y cupula frontal.",
        "Colocar provisionalmente los elementos para medir longitud de cables.",
        "Instalar en subchasis asegurando 100% horizontalidad.",
        "Usar velcro industrial o silentblocks para proteger el disco interno.",
        "Conectar a 12V bajo llave y asegurar buena masa.",
        "Fabricar/adaptar pletina de aluminio o carbono.",
        "Guiar cable rizado sin que interfiera con el giro del manillar.",
        "Pegar con cinta 3M de doble cara en colin o cupula.",
        "Comprobar vision directa al cielo sin obstrucciones.",
        "Instalar abrazaderas en botella y soporte de pinza.",
        "Sensor debe trabajar completamente paralelo a la horquilla.",
        "Comprimir horquilla a tope; el sensor NO debe llegar a su limite.",
        "Fabricar pletina en pinza de freno (trasera pref.).",
        "Ajustar distancia a 1.0 - 2.0 mm con galga.",
        "Pasar cables pegados a vigas principales con bridas.",
        "Distancia >5cm de colectores y libre de puntos de giro.",
        "Conectar CAN Bus a OBD o cable inductivo a bobina.",
        "Verificar que EVO4 y Dash encienden correctamente.",
        "Abrir AiM RaceStudio en el portatil.",
        "Configurar modelo de moto y protocolo CAN.",
        "Calibrar punto Cero con la moto extendida en caballetes.",
        "Girar rueda manualmente y verificar lectura de velocidad.",
    ],
    "Check": ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
}

# Datos para la Pestana 4: Materiales y Herrajes
data4 = {
    "Categoria": [
        "Soportes",
        "Soportes",
        "Tornilleria",
        "Fijacion",
        "Fijacion",
        "Proteccion",
        "Proteccion",
        "Gestion",
        "Quimicos",
        "Quimicos",
        "Electrico",
    ],
    "Material / Pieza": [
        "Pletina de Aluminio",
        "Abrazaderas de Horquilla",
        "Kit Tornilleria Inox",
        "Velcro Industrial",
        "Adhesivo Doble Cara",
        "Cinta de Tela (Tesa)",
        "Tubo Termorretractil",
        "Bridas de Nylon",
        "Fijador de Roscas",
        "Limpiador Superficies",
        "Porta-fusible aereo",
    ],
    "Especificacion Tecnica": [
        "Grado 6061-T6 (2mm)",
        "Anillos mecanizados a medida",
        "Acero Inox A2 (M4, M5, M6)",
        "3M Dual Lock",
        "3M VHB (Very High Bond)",
        "Estilo OEM (Motorcycle)",
        "Con adhesivo interno",
        "Resistentes UV/Calor",
        "Loctite 243 (Azul)",
        "Alcohol Isopropilico",
        "Mini-blade (Fusible 5A)",
    ],
    "Cantidad Est.": ["1 unidad", "2 unidades", "1 kit", "20 cm", "1 rollo", "1 rollo", "1 kit", "50 unidades", "1 bote", "1 bote", "1 unidad"],
    "Proposito de Instalacion": [
        "Soportes de Dash y Sensor Velocidad.",
        "Anclaje de potenciometro a suspension.",
        "Fijacion de herrajes y sensores.",
        "Sujecion de EVO4 y Dash.",
        "Fijacion de antena GPS.",
        "Proteccion mazo de cables contra roces.",
        "Sellado estanco de conexiones.",
        "Organizacion del cableado por el chasis.",
        "Seguridad contra vibraciones del motor.",
        "Limpieza previa de zonas de contacto.",
        "Proteccion linea de alimentacion.",
    ],
}

# Datos para futura impresion 3D (Pestana 5)
data5 = {
    "Pieza 3D": [
        "Soporte Dash frontal",
        "Soporte antena GPS",
        "Soporte sensor velocidad",
        "Guia de cableado lateral",
        "Abrazadera potenciometro",
    ],
    "Zona de montaje": [
        "Arana frontal",
        "Colin / cupula",
        "Pinza freno / basculante",
        "Vigas chasis",
        "Horquilla delantera",
    ],
    "Objetivo funcional": [
        "Reducir vibracion y facilitar lectura",
        "Mejor vista al cielo y fijacion limpia",
        "Mantener gap estable 1.0-2.0 mm",
        "Ordenar mazo y evitar rozaduras",
        "Alineacion paralela y ajuste fino",
    ],
    "Material recomendado": ["PA12", "ASA", "Nylon CF", "PETG", "Nylon CF"],
    "Proceso sugerido": ["SLS", "FDM", "SLS", "FDM", "SLS"],
    "Tolerancia objetivo": ["+/-0.20 mm", "+/-0.30 mm", "+/-0.20 mm", "+/-0.30 mm", "+/-0.15 mm"],
    "Estado CAD": ["Pendiente", "Pendiente", "Pendiente", "Pendiente", "Pendiente"],
    "Archivo STL": ["", "", "", "", ""],
    "Notas de impresion": [
        "Insertos metalicos para tornilleria M4/M5",
        "Resistencia UV y temperatura",
        "Alta rigidez para no perder calibracion",
        "Radio minimo en cantos para evitar rotura",
        "Validar recorrido completo de suspension",
    ],
}


def autosize_columns(writer, sheet_name, dataframe):
    worksheet = writer.sheets[sheet_name]
    for idx, col in enumerate(dataframe.columns, start=1):
        max_len = max(dataframe[col].astype(str).map(len).max(), len(col))
        col_letter = get_column_letter(idx)
        worksheet.column_dimensions[col_letter].width = min(max_len + 2, 60)


def style_worksheet(writer, sheet_name, dataframe):
    worksheet = writer.sheets[sheet_name]
    max_row = len(dataframe) + 1
    max_col = len(dataframe.columns)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = thin_border
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    worksheet.print_title_rows = "1:1"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0


def normalize_text(value):
    replacements = {
        "a": "a",
        "e": "e",
        "i": "i",
        "o": "o",
        "u": "u",
        "n": "n",
    }
    text = str(value).lower()
    for original, clean in {"á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u", "ñ": "n"}.items():
        text = text.replace(original, clean)
    text = text.replace("_", " ").replace("-", " ")
    return " ".join(text.split())


def find_stl_files(base_dir):
    base_path = Path(base_dir)
    matches = list(base_path.rglob("*.stl")) + list(base_path.rglob("*.STL"))
    rows = []

    for file_path in sorted(matches):
        relative_path = file_path.relative_to(base_path)
        rows.append(
            {
                "Archivo": file_path.name,
                "Ruta relativa": str(relative_path),
                "Tamano (KB)": round(file_path.stat().st_size / 1024, 2),
            }
        )

    return rows


def enrich_3d_plan_with_detected_stl(df_3d_plan, stl_rows):
    if not stl_rows:
        return df_3d_plan

    search_pool = [
        {
            "name": row["Archivo"],
            "path": row["Ruta relativa"],
            "norm": normalize_text(row["Archivo"]),
        }
        for row in stl_rows
    ]

    def build_keywords(piece_name):
        ignore = {"soporte", "pieza", "frontal", "de", "del", "la", "el", "y"}
        tokens = [token for token in normalize_text(piece_name).split() if len(token) > 2 and token not in ignore]
        return tokens[:4]

    df_copy = df_3d_plan.copy()

    for idx, row in df_copy.iterrows():
        keywords = build_keywords(row["Pieza 3D"])
        best_match = None
        best_score = 0

        for candidate in search_pool:
            score = sum(1 for keyword in keywords if keyword in candidate["norm"])
            if score > best_score:
                best_score = score
                best_match = candidate

        if best_match and best_score >= 1:
            df_copy.at[idx, "Archivo STL"] = best_match["path"]
            df_copy.at[idx, "Estado CAD"] = "STL detectado"

    return df_copy


def create_bom_dataframe(materials_df):
    bom_df = materials_df.copy()
    bom_df["Proveedor sugerido"] = "Pendiente"
    bom_df["Referencia proveedor"] = "Pendiente"
    bom_df["Coste estimado (EUR)"] = 0.0
    bom_df["Coste total (EUR)"] = 0.0
    return bom_df


def main():
    df1 = pd.DataFrame(data1)
    df2 = pd.DataFrame(data2)
    df3 = pd.DataFrame(data3)
    df4 = pd.DataFrame(data4)
    df5 = pd.DataFrame(data5)
    df6 = create_bom_dataframe(df4)

    stl_rows = find_stl_files(Path(__file__).resolve().parent)
    df5 = enrich_3d_plan_with_detected_stl(df5, stl_rows)
    if stl_rows:
        df7 = pd.DataFrame(stl_rows)
    else:
        df7 = pd.DataFrame(
            [
                {
                    "Archivo": "Sin STL detectados",
                    "Ruta relativa": "-",
                    "Tamano (KB)": 0,
                }
            ]
        )

    file_path = "Plan_Instalacion_Telemetria_AiM_CFMoto675SR.xlsx"

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df1.to_excel(writer, sheet_name="01_Especificaciones", index=False)
        df2.to_excel(writer, sheet_name="02_Protocolo_SOP", index=False)
        df3.to_excel(writer, sheet_name="03_Checklist", index=False)
        df4.to_excel(writer, sheet_name="04_Materiales", index=False)
        df5.to_excel(writer, sheet_name="05_Piezas_3D_Futuro", index=False)
        df6.to_excel(writer, sheet_name="06_BOM_Compras", index=False)
        df7.to_excel(writer, sheet_name="07_STL_Encontrados", index=False)

        autosize_columns(writer, "01_Especificaciones", df1)
        autosize_columns(writer, "02_Protocolo_SOP", df2)
        autosize_columns(writer, "03_Checklist", df3)
        autosize_columns(writer, "04_Materiales", df4)
        autosize_columns(writer, "05_Piezas_3D_Futuro", df5)
        autosize_columns(writer, "06_BOM_Compras", df6)
        autosize_columns(writer, "07_STL_Encontrados", df7)

        style_worksheet(writer, "01_Especificaciones", df1)
        style_worksheet(writer, "02_Protocolo_SOP", df2)
        style_worksheet(writer, "03_Checklist", df3)
        style_worksheet(writer, "04_Materiales", df4)
        style_worksheet(writer, "05_Piezas_3D_Futuro", df5)
        style_worksheet(writer, "06_BOM_Compras", df6)
        style_worksheet(writer, "07_STL_Encontrados", df7)

    print(f"Excel generado: {file_path}")


if __name__ == "__main__":
    main()
